import time
import numpy as np
from openpyxl import Workbook

"""
a – кол-во правильно определенных  классификатором документов для категории
с - кол-во неправильно определенных классификатором для категории товаров =m-a
d - кол-во, которое не нашел, а должен был найти для категории
l - всего товар категории д.б.б. найдено =a+d
m - сколько всего записал в категорию товаров

R - полнота. R= a/(а+c), где a – кол-во правильно определенных  классификатором документов и с кол-во документов, которое классификатор не определил для классификатора.
Полнота – доля документов действительно принадлежащих к данному классу, которую нашел классификатор от всего объема документов, которые действительно принадлежат классу.
сколько нашел от того, что должен был найти для класса  

P - точность P=a/(a+b), где а - кол-во правильно определенных и b -кол-во неправильно определенных документов к одному классу.
Точность – доля правильно определенных документов данного класса от всего кол-во документов, которое классификатор записал в данный класс.
Из всего что записал в класс, насколько правильно 

F-мера - содержит баланс межлу R и P  F=2pr/p+r
Для общего используем макроусреднение - составляем по каждому и делим на кол-во

В качестве маркера для проверки базового значения проверяем на тех же данных, что и при обучении, чтобы исключить наличие ошибок классификации уже на этом этапе  

"""


def quality_classification(goods_description, goods_classes_by_teacher, goods_classes_by_algorithm, start_time_edu,
                           start_time_clas, goods_supplier_classes):
    # считаем время
    general_time = timer(start_time_edu)
    classific_time = timer(start_time_clas)
    education_time = start_time_clas - start_time_edu
    goods_quantity = len(goods_description)

    # считаем общее кол-во корректно и ошибок (E) + вычисляем  a, c, d
    correct_estimates = int(0)  # общее кол-во правильно категоризированных товаров
    error_estimates = int(0)  # общее кол-во ошибочно категоризированных товаров
    list_categories = set(goods_classes_by_teacher)  # получаем список всех категорий, которые были в тестовой выборке

    dict_a = dict.fromkeys(list_categories, 0)  # сколько истинно правильно нашли для каждой категории -a
    dict_d = dict.fromkeys(list_categories, 0)  # сколько должен был найти, но не нашел для категорий - d
    #print(goods_classes_by_algorithm)

    for i in range(len(goods_classes_by_teacher)):
        # goods_classes_by_teacher[i]="'"+str(goods_classes_by_teacher[i])+"'"
        #print(goods_classes_by_teacher[i], 'ftf')
        tp = goods_classes_by_teacher[i]
        tr = goods_classes_by_algorithm[i]
        #print(goods_classes_by_algorithm[i], 'ara')
        if goods_classes_by_teacher[i] == goods_classes_by_algorithm[i]:
            correct_estimates = correct_estimates + 1  # обновляем общее число правильно определенных категорий
            dict_a[goods_classes_by_teacher[i]] = dict_a.get(goods_classes_by_teacher[i]) + 1  # считаем а
        else:
            dict_d[goods_classes_by_teacher[i]] = dict_d.get(goods_classes_by_teacher[i]) + 1  # считаем d
            error_estimates = error_estimates + 1  # обновляем общее число ошибочно определенных категорий

    dict_m = {}  # кол-во m (сколько всего было записано) для каждой категрии от алгоритма
    for value in list_categories:
        quantity = np.ndarray.tolist(goods_classes_by_algorithm).count(value)
        dict_m[value] = quantity

    dict_c = {}  # кол-во ошибочно определенных для категории товаров
    for value in list_categories:
        dict_c[value] = dict_m[value] - dict_a[value]  # для каждой категории из m-a

    # считаем P - точность a/(a+c) и R - полнота a/(a+d)
    dict_p = {}  # P - точность
    dict_r = {}  # R - полнота
    dict_e = {}

    for value in list_categories:
        dict_p[value] = dict_a[value] / (dict_a[value] + dict_c[value])
        dict_r[value] = dict_a[value] / (dict_a[value] + dict_d[value])
        dict_e[value] = (dict_c[value] + dict_d[value]) / (
                dict_c[value] + dict_d[value] + dict_a[value] + dict_a[value])

    dict_f = {}  # F-мера
    for value in list_categories:
        dict_f[value] = (2 * dict_p[value] * dict_r[value]) / (dict_p[value] + dict_r[value])

    f_general = (sum(dict_f.values())) / len(dict_f)  # F-мера средняя по всем категориям

    # сохраняем значения в excel
    wb = Workbook()
    path_to_file = 'file_to_load/output.xlsx'

    # заполняем первую вкладку значениями классификатора
    ##заголовок таблицы
    ws1 = wb.active
    ws1.title = 'результат классификации'
    column_titles = {0: '№ товара', 1: 'Путь классификатора поставщика', 2: 'ID классификатора учителя',
                     3: 'ID классификатора алогоритма'}
    ##значения таблицы классификатора
    counter = int(0)
    while counter < len(goods_supplier_classes) + 1:
        if counter == 0:
            for col in range(1, 5):
                ws1.cell(column=col, row=counter + 1, value='{0}'.format(column_titles[col - 1]))
        else:
            for col in range(1, 5):
                if col == 1:
                    ws1.cell(column=col, row=counter + 1, value='{0}'.format(counter))
                elif col == 2:
                    ws1.cell(column=col, row=counter + 1, value='{0}'.format(goods_supplier_classes[counter - 1]))
                elif col == 3:
                    ws1.cell(column=col, row=counter + 1, value='{0}'.format(goods_classes_by_teacher[counter - 1]))
                else:
                    ws1.cell(column=col, row=counter + 1, value='{0}'.format(goods_classes_by_algorithm[counter - 1]))
        counter = counter + 1
    # заполняем вторую вкладку результаты анализа
    ws2 = wb.create_sheet(title="Оценка классификации")
    ##заполняем заголовки общей таблицы
    column_titles2 = {0: '№', 1: 'Класс', 2: 'P-точность', 3: 'R-Полнота',
                      4: 'E-Ошибка', 5: 'F-мера'}

    for col in range(1, 7):
        ws2.cell(column=col, row=1, value='{0}'.format(column_titles2[col - 1]))
    ##заполяняем значения таблицы
    row_no = 2
    for value in list_categories:
        for col in range(1, 7):
            if col == 1:
                ws2.cell(column=col, row=row_no, value='{0}'.format(row_no - 1))
            elif col == 2:
                ws2.cell(column=col, row=row_no, value='{0}'.format(value))
            elif col == 3:
                ws2.cell(column=col, row=row_no, value='{0}'.format(dict_p[value]))
            elif col == 4:
                ws2.cell(column=col, row=row_no, value='{0}'.format(dict_r[value]))
            elif col == 5:
                ws2.cell(column=col, row=row_no, value='{0}'.format(dict_e[value]))
            else:
                ws2.cell(column=col, row=row_no, value='{0}'.format(dict_f[value]))
        row_no = row_no + 1

    ##заполняем заголовки футтера таблицы доп значениями
    column_titles3 = {0: 'Общее время:', 1: 'Всего классифицировано:', 2: 'Из них правильно',
                      3: 'Из них не правильно', 4: 'Общая сбалансированная  F'}  # добавить кол-во ошибок на категорию
    output_values = {0: general_time, 1: goods_quantity, 2: correct_estimates,
                     3: error_estimates, 4: f_general}

    ##заполняем значения футтера таблицы доп значениями
    i = 0
    for row in range(row_no, row_no + 5, 1):
        for col in range(1, 2):
            ws2.cell(column=col, row=row, value='{0}'.format(column_titles3[i]))
        for col in range(2, 3):
            ws2.cell(column=col, row=row, value='{0}'.format(output_values[i]))
        i = i + 1

    wb.save(filename=path_to_file)

    return 'Оценка классификации:', \
           '\nОбщее время выполнения оценки: {0} сек. Из них:'.format(general_time), \
           '\t На обучение:{0} сек.'.format(education_time), \
           '\t На классификацию:{0} сек.'.format(classific_time), \
           '\nВсего было классифицировано: {0} тов. Из них:'.format(goods_quantity), \
           '\t Правильно классифицированы:{0} тов.'.format(correct_estimates), \
           '\t Ошибочно классифицированы:{0} тов.'.format(error_estimates), \
           '\nКачество классификации по категориям:', \
           '\tТочность по категория:\n\t\t{0}'.format(dict_p), \
           '\tПолнота по категориям:\n\t\t{0}'.format(dict_r), \
           '\tСбалансированная F1 мара:\n\t\t{0}'.format(dict_f), \
           '\nДетальная информация сохранена в файл.'


# start_time = time.time()
def timer(start):
    # time_load_for = "Загрузка выборки заняла {0} сек."
    time_result = time.time() - start
    # time_exect = time_load_for.format(time_result)
    return time_result
