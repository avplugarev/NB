import openpyxl
from openpyxl import Workbook
from pymorphy2 import MorphAnalyzer
import bd_connector as bd
from sklearn.feature_extraction.text import TfidfVectorizer

"""
  Класс загрузки/выгрузки данных в/из системы для дальнейшей его обработки
"""


class DataLoader():
    path = str()

    def __init__(self):
        self.goods_description = []  # пустой список под описагния товаров
        self.goods_target_classes = []  # пустой список под категории для товаров
        self.goods_supplier_classes = []  # пустой список ппод классификатор поставщика
        self.raw_description = list()

    """
        метод загрузки данных из файла
    """

    def import_file_data(self, path):
        file = openpyxl.load_workbook(path)  # загружам файл в объект для дальнейшей работе с ним
        working_sheet = file.active  # получаю объект лист для работы
        # чистим выгрузку от лишних ячеек
        working_sheet = DataPreprocessed.clear_file_data(working_sheet)
        for row in working_sheet.values:
            # добавляем категорию товара в коллекцию с таргетинговыми значениями
            self.goods_target_classes.append(row[0])
            # производим работы с описанием товара и добавляем его в обучающую или тестовую выгрузку
            good = str('')
            goods_supplier_class = str()
            n = 0
            for value in row[1:]:
                # получаем путь к категории классификатора поставщика для этого товара
                if n < 6:
                    goods_supplier_class = DataPreprocessed.sup_class_replacement(goods_supplier_class, value)
                n = n + 1
                # получаем обработанное описание товара для дальнейшей векторизации
                good = DataPreprocessed.text_preprocessed(value, row, good)
            print("Извлечен торва №{0}: {1}".format(len(self.goods_description), good))
            # добавляем описание товара в коллекцию описания товаров
            self.goods_description.append(str.strip(good))
            # добавляем описание пути к категории в коллекции пставщика
            self.goods_supplier_classes.append(str.strip(goods_supplier_class))
        file.close()
        return self

    """
        метод сохраняния обучающей выборки в бд
    """
    def load_educ_data_to_bd(data):
        print(bd.edu_data_delete())  # чистим базу данных для обучения
        educt_data_counter = bd.read_prepared_data()[2]  # получаем id последнего добавленного товара
        for i in range(len(data.goods_description)):
            data_to_add_bd = dict()
            educt_data_counter = educt_data_counter + 1
            data_to_add_bd['uid_educ'] = educt_data_counter
            data_to_add_bd['description'] = data.goods_description[i]  # описание товара из списка для обучения
            data_to_add_bd['category'] = data.goods_target_classes[i]  # значение категории для товара для обучения
            data_to_add_bd['path'] = data.goods_supplier_classes[i]  # путь к категории поставщика
            bd.add_good_to_prepared_data(data_to_add_bd)  # сохраняем в базу потоварно
        return 'Обучающие данные загружены в базу'

    """
        метод загрузки классификатора купивип в базу
    """

    def import_classifier(path):
        file = openpyxl.load_workbook(path)  # загружам файл в объект для дальнейшей работе с ним
        working_sheet = file.active  # получаю объект лист ддя работы
        print(bd.classifier_kupivip_delete())

        for row in working_sheet.values:
            category_id = str()  # поле id категории
            category_path = str()  # поле путь классификатора к категории
            data = {}  # словарь записываемого значения в бд

            for value in row[:1]:
                category_id = value
            for value in row[1:7]:
                if value == None:
                    category_path = category_path + ' ' + str('None')
                else:
                    category_path = category_path + ' ' + value
            data['category_id'] = category_id
            data['path_category'] = category_path
            bd.add_category_kupivip(data)
        file.close()
        return 'Данные по категориям успешно загружены'

    """
        TODO: метод загрузки данных через API
    """

    def import_api_data(data):
        api_request = DataLoader()
        api_request.raw_description.append(data)#  сохранили изначальное значение
        api_request.goods_description = DataPreprocessed.get_correct_fields(data)  # оставили корректные поля
        # получили путь к классификатору поставщика
        api_request.goods_supplier_classes = DataPreprocessed.get_suplier_classifier(api_request.goods_description)

        # заменяем значениями из справочников
        api_request.goods_description = DataPreprocessed.enriched_fields(api_request.goods_description)
        api_request.goods_description = DataPreprocessed.remove_empty_fields(api_request.goods_description)

        return api_request


    """
        метод выгрузки данных по статистики в Excel
    """

    def export_file_data(path, goods, goods_classes_by_algorithm, estimates):
        pass
        # сохраняем значения в excel
        wb = Workbook()
        list_categories = set(goods.goods_target_classes)  # получаем список всех категорий в тестовой выборке

        # заполняем первую вкладку значениями классификатора
        # заголовок таблицы
        ws1 = wb.active
        ws1.title = 'результат классификации'
        column_titles = {0: '№ товара',
                         1: 'Путь классификатора поставщика',
                         2: 'ID классификатора учителя',
                         3: 'ID классификатора алогоритма',
                         4: 'Путь классификатора купивип'}
        # значения таблицы классификатора
        counter = int(0)
        while counter < len(goods.goods_supplier_classes) + 1:
            if counter == 0:
                for col in range(1, 6):
                    ws1.cell(column=col, row=counter + 1, value='{0}'.format(column_titles[col - 1]))
            else:
                for col in range(1, 6):
                    if col == 1:
                        ws1.cell(column=col, row=counter + 1, value='{0}'.format(counter))
                    elif col == 2:
                        ws1.cell(column=col, row=counter + 1,
                                 value='{0}'.format(goods.goods_supplier_classes[counter - 1]))
                    elif col == 3:
                        ws1.cell(column=col, row=counter + 1,
                                 value='{0}'.format(goods.goods_target_classes[counter - 1]))
                    elif col == 4:
                        ws1.cell(column=col, row=counter + 1,
                                 value='{0}'.format(goods_classes_by_algorithm[counter - 1]))
                    else:
                        category_path_kupivip = bd.get_category_kupivip_by_id(
                            goods_classes_by_algorithm[counter - 1])
                        ws1.cell(column=col, row=counter + 1, value='{0}'.format(category_path_kupivip))
            counter = counter + 1
        # заполняем вторую вкладку результаты анализа
        ws2 = wb.create_sheet(title="Оценка классификации")
        # заполняем заголовки общей таблицы
        column_titles2 = {0: '№', 1: 'Класс', 2: 'P-точность', 3: 'R-Полнота',
                          4: 'E-Ошибка', 5: 'F-мера'}

        for col in range(1, 7):
            ws2.cell(column=col, row=1, value='{0}'.format(column_titles2[col - 1]))
        # заполяняем значения таблицы
        row_no = 2
        for value in list_categories:
            for col in range(1, 7):
                if col == 1:
                    ws2.cell(column=col, row=row_no, value='{0}'.format(row_no - 1))
                elif col == 2:
                    ws2.cell(column=col, row=row_no, value='{0}'.format(value))
                elif col == 3:
                    ws2.cell(column=col, row=row_no, value='{0}'.format(estimates.dict_p[value]))
                elif col == 4:
                    ws2.cell(column=col, row=row_no, value='{0}'.format(estimates.dict_r[value]))
                elif col == 5:
                    ws2.cell(column=col, row=row_no, value='{0}'.format(estimates.dict_e[value]))
                else:
                    ws2.cell(column=col, row=row_no, value='{0}'.format(estimates.dict_f[value]))
            row_no = row_no + 1

        # заполняем заголовки футтера таблицы доп значениями
        column_titles3 = {0: 'Общее время:', 1: 'Всего классифицировано:', 2: 'Из них правильно',
                          3: 'Из них не правильно',
                          4: 'Общая сбалансированная  F'}  # добавить кол-во ошибок на категорию
        output_values = {0: estimates.general_time, 1: estimates.goods_quantity, 2: estimates.correct_estimates,
                         3: estimates.error_estimates, 4: estimates.f_general}

        # заполняем значения футтера таблицы доп значениями
        i = 0
        for row in range(row_no, row_no + 5, 1):
            for col in range(1, 2):
                ws2.cell(column=col, row=row, value='{0}'.format(column_titles3[i]))
            for col in range(2, 3):
                ws2.cell(column=col, row=row, value='{0}'.format(output_values[i]))
            i = i + 1

        wb.save(filename=path)


"""
    Класс обработки загруженных данных для дальнейшей оценки
"""


class DataPreprocessed():

    def __init__(self):
        self.good_rows = [
            'Class_ID',
            'Уровень 1',
            'Уровень 2',
            'Уровень 3',
            'Уровень 4',
            'Уровень 5',
            'Уровень 6',
            'Наименование (<=250)',
            'Техническое описание',
            'Пол',
            'Размер производителя  (<=10)',
            'НДС',
            'Возраст+',
            'Тип чипа',
            'Тип ткани',
            'Тип кроя/джинсы'
        ]
        self.gender = {
            0: 'любой_пол',
            1: 'мужчина',
            2: 'женщина'
        }
        self.vat = {
            0: 'медицина',
            10: 'детский',
            20: 'товар_с_обычным_налогом'
        }
        self.trackingType = {
            'DMX': 'обувь',
            'datamatrix': 'обувь',
            'MEX': 'мех',
            'FUR': 'мех',
            'LOT': 'без_чипа',
            'обычный': 'без_чипа',
            'Обычный': 'без_чипа',
            'обычно': 'без_чипа',
            'LPR': 'легпром',
            'легпром': 'легпром',
            'PRF': 'парфюм',
            'парфюм': 'парфюм'
        }
        self.morph = MorphAnalyzer()
        self.stop_symbl = ['.', ',', '!', ':', ';', ')', '(', '«', '»', '"']
        self.api_good_allowed_inputs = [
            'level1',
            'level2',
            'level3',
            'level4',
            'level5',
            'level6',
            'name',
            'composition',
            'technicalDescription',
            'gender',
            'items',
            'vat',
            'age',
            'trackingType',
            'material',
            'jeansCut'
        ]
        self.api_dicts = [
            'gender',
            'vat',
            'trackingType',
            'material'
        ]
        self.material = {
            1: 'текстиль',
            2: 'трикотаж',
            3: 'другое'
        }

    """ 
        метод удаляения пустых и не нужных нам столбцов имеющих шумовой эффект из файла
    """

    def clear_file_data(working_sheet):
        dicts_for_data = DataPreprocessed()
        for row in working_sheet:
            for cell in row:
                if cell.value not in dicts_for_data.good_rows:
                    working_sheet.delete_cols(cell.column)  # удаляем столбцы, если их заголовки нам не нужны
            break
        working_sheet.delete_rows(0)
        return working_sheet

    """ 
        метод удаляения не нужных полей из запроса API
    """

    def get_correct_fields(data):
        self = DataPreprocessed()
        data_to_return = list()

        for good in data:
            good_desc = dict()
            for i in range(len(self.api_good_allowed_inputs)):
                if i <6 and good.get(self.api_good_allowed_inputs[i], 'None')=='' or i <6 and good.get(self.api_good_allowed_inputs[i], 'None')=='None':
                    good_desc[self.api_good_allowed_inputs[i]]='None'
                else:
                    if good.get(self.api_good_allowed_inputs[i], 'None') != 'None':
                        good_desc[self.api_good_allowed_inputs[i]]=good.get(self.api_good_allowed_inputs[i], 'None')
            good_desc['items'] = good_desc['items'][0].get('variantCode', 'None')
            data_to_return.append(good_desc)
        return data_to_return

    """ 
        метод замены данных по справочникам через API и удаления шумовых данных с нормализацией
    """
    def enriched_fields(description):
        self = DataPreprocessed()
        for value in description:
            for field in value:
                word=str()
                if field == 'gender':
                    value['gender']=self.gender.get(value[field])
                elif field == 'vat':
                    value['vat'] = self.vat.get(value[field])
                elif field == 'trackingType':
                    value['trackingType'] = self.trackingType.get(value[field])
                elif field == 'material':
                    value['material'] = self.material.get(value[field])
                else:
                    value[field]=str.strip(DataPreprocessed.word_normalisation(self,value.get(field),word))
        return description

    """ 
        метод очистки пустых полей для тов через API
    """
    def remove_empty_fields(goods):
        self = DataPreprocessed()
        ready_to_classify=list()
        for good in goods:
            good_description = str()
            for i in range(len(self.api_good_allowed_inputs)):
                if good.get(self.api_good_allowed_inputs[i])!=None \
                        and good.get(self.api_good_allowed_inputs[i])!=''\
                        and good.get(self.api_good_allowed_inputs[i])!='None':
                    good_description=good_description + str(good.get(self.api_good_allowed_inputs[i])) +' '
            ready_to_classify.append(str.strip(good_description))
        return ready_to_classify

    """ 
            метод получения пути через API
    """

    def get_suplier_classifier(data):
        list_of_pathes = list()
        for good in data:
            path = list()
            for value in good:
                if len(path) < 6:
                    if good.get(value, None) == '' or good.get(value, None) == 'None':
                        path.append('None')
                    else:
                        path.append(good.get(value, None))
                else:
                    break
            list_of_pathes.append(path)
        return list_of_pathes

    """
        метод замены пустых значений для пути классификатора поставщика на None
    """

    def sup_class_replacement(goods_supplier_class, value):
        if value != None:
            goods_supplier_class = goods_supplier_class + str(value) + ' '
        else:
            goods_supplier_class = goods_supplier_class + str('None') + ' '
        return goods_supplier_class

    """
        метод обогащения числовых значений текстовыми, удаления спецсимволов  
    """

    def text_preprocessed(value, row, good):
        self = DataPreprocessed()
        # удаляем пустые ячейки
        if value == None:
            return good
        # заменяем числовые значения в описание товара - значениями из справочника
        elif value == row[9]:
            return str.strip(good) + ' ' + self.gender.get(value, None)
        elif value == row[11]:
            return str.strip(good) + ' ' + self.vat.get(value, None)
        elif value == row[13]:
            return str.strip(good) + ' ' + self.trackingType.get(value, None)
        elif value == row[14]:
            try:
                return str.strip(good) + ' ' + str(self.material.get(value, None))
            except:
                print(str.strip(good),'+',value,'+',self.material.get(value, None) )
        else:
            return DataPreprocessed.word_normalisation(self, value, good)

    """
         метод приведения слов к нормальной форме, удаление лишних форм речи
    """

    def word_normalisation(self, value, good):
        new_str = str()
        value = str.strip(str(value))  # убираем пробелы
        #  удаляем спецсимволы
        for i in range(len(self.stop_symbl)):
            value = value.replace(self.stop_symbl[i], '')
        value = value.split()
        # приводим к нормальной форме и оставляем только сущ и прилагательные в ед числе
        for word in value:
            word = str.strip(word)
            word = self.morph.parse(word)[0].normal_form
            if self.morph.parse(word)[0].tag.POS == 'NOUN' or \
                    self.morph.parse(word)[0].tag.POS == 'ADJF':
                if self.morph.parse(word)[0].inflect({'sing', 'nomn'}) == None:
                    word = self.morph.parse(word)[0].normal_form
                    new_str = ''.join(word)
                else:
                    word = self.morph.parse(word)[0].inflect({'sing', 'nomn'}).word
                    new_str = ''.join(word)
        return good + ' ' + new_str

    """
        метод перевода описания товара в векторную форму
    """

    def get_vocabular(model_goods_description):
        print('Начали формировать словарь')
        model_vocabular = TfidfVectorizer()
        model_vocabular.fit(model_goods_description)  # заполняем словарь и токенизируем
        print('Сформировали словарь')
        return model_vocabular

    """ 
        метод перевода описаний товаров в вектор
    """

    def goods_vectors(data, model_vocabular):
        good_without_class_vector = model_vocabular.transform(data)  # переводим описание в векторную форму
        return good_without_class_vector.toarray()

# path = 'file_to_load/teach_file_sample.xlsx'

# data = DataLoader()
# data = DataLoader.import_file_data(data, path)
# сохраняем обучающую выборку в базу
# DataLoader.load_educ_data_to_bd(data)
# print(data.goods_description)
# print(data.goods_target_classes)
# print(data.goods_supplier_classes)
#path_to_classifier = 'file_to_load/class_ready.xlsx'
#print(DataLoader.import_classifier(path_to_classifier))
